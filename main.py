import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
import shutil
import datetime
import os

def main():
    # Hardcoded paths (relative to project root)
    template_path = 'templates/template.xlsx'  # Your template file
    pickle_path = 'pickles/data.pkl'          # Your pickle file
    
    # Generate output filename with date prefix (e.g., 20240209_template.xlsx)
    today = datetime.date.today().strftime('%Y%m%d')
    output_filename = f"{today}_template.xlsx"
    output_path = os.path.join('output', output_filename)
    
    # Copy template to output folder
    shutil.copyfile(template_path, output_path)
    print(f"Template copied to {output_path}")
    
    # Load Pandas DataFrame from pickle
    df = pd.read_pickle(pickle_path)
    df['uuid'] = df['uuid'].astype(str)

    # Calculate column 'Duracion'
    df['Duracion'] = df.apply(
        lambda row: (pd.to_datetime(row['FinEvento']) - pd.to_datetime(row['InicioEvento'])).total_seconds() / 60,
        axis=1
    )
    df['Duracion'] = df['Duracion'].round(0).astype(int).astype(str)
    cols = df.columns.tolist()
    duracion_index = cols.index('Duracion')  # Find where we temporarily added it
    cols.insert(15, cols.pop(duracion_index))  # Move it to position 14
    df = df[cols]

    print(f"Loaded DataFrame with {len(df)} rows")
    
    # Load the copied workbook
    wb = load_workbook(output_path)
    
    # Access the second sheet (0-based index; change if needed)
    sheet = wb.worksheets[1]  # Or wb['Sheet2'] if named
    
    # Optional: Clear existing data from row 2 down (preserves headers and formats)
    for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=1, max_col=sheet.max_column):
        for cell in row:
            cell.value = None
    
    # Insert DataFrame starting from row 2 (skip headers in DF)
    for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=False), 2):
        for c_idx, value in enumerate(row, 1):
            sheet.cell(row=r_idx, column=c_idx, value=value)
    
    # Save the modified workbook
    wb.save(output_path)
    print(f"Data inserted into {output_path}")

if __name__ == "__main__":
    main()
